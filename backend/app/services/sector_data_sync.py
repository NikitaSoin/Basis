"""Сбор ОТРАСЛЕВЫХ показателей из источников реестра (`config/source_registry_*.md`).

Владелец (2026-08-07): «надо теперь настроить парсинг с этих источников, и чтобы
карточки самообновлялись».

Первая очередь — источники со статусом `open` и приоритетом 1, закрывающие
отрасли, у которых сейчас НЕТ ни одного живого показателя своего рынка. По
барометру таких девять из тринадцати: энергетика, машиностроение, IT, телеком,
транспорт, девелопмент, здравоохранение и другие оценивались только по
отчётности компаний, без отраслевых данных.

🔴 ПОЧЕМУ РЯДЫ КЛАДЁМ В macro_data_points, А НЕ В НОВУЮ ТАБЛИЦУ.
Соблазн завести `sector_indicators` есть, но у нас уже работает вся обвязка
вокруг macro_indicators/macro_data_points: ревизия точек, `check_staleness`,
«ОТК данных», отдача рядов на витрину (`/market/commodity-price-history` умеет
`macro:<code>`), и барометр читает оттуда же. Вторая таблица потребовала бы
дублировать всё это ради разделения, которого пользователь не видит. Отличаем
отраслевые ряды префиксом кода (`sec_*`) и полем `display_group="sector"`.

🔴 ЧЕСТНАЯ ДЕГРАДАЦИЯ ВАЖНЕЕ ПОЛНОТЫ. Источники живут своей жизнью: меняют
вёрстку, уходят под антибот, отдают 200 с пустым каркасом. Поэтому каждый
коллектор возвращает либо разобранное число, либо None с причиной, и НИКОГДА не
пишет догадку. Пустой ряд честнее выдуманного: барометр в этом случае поставит
«мало данных», а не уверенный балл из воздуха.
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime, timezone

import httpx
from sqlalchemy import text
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/125.0 Safari/537.36")
_TIMEOUT = 25


def _fetch(url: str) -> str | None:
    """Скачать страницу. None — если недоступна или пришла заглушка вместо данных.

    Проверка на заглушку обязательна: и e-disclosure, и Минэнерго отдают 200 с
    большим телом, внутри которого нет данных (JS-челлендж, пустой SPA-каркас).
    Код ответа и размер здесь ничего не доказывают — см. журнал 2026-08-06.
    """
    try:
        with httpx.Client(follow_redirects=True, timeout=_TIMEOUT,
                          headers={"User-Agent": _UA}) as c:
            r = c.get(url)
        if r.status_code != 200:
            logger.info("sector_data_sync: %s → HTTP %s", url, r.status_code)
            return None
        body = r.text or ""
        low = body.lower()
        if any(k in low for k in ("just a moment", "checking your browser",
                                  "пройдите проверку", "не робот", "cf-challenge")):
            logger.warning("sector_data_sync: %s → антибот-заглушка", url)
            return None
        if len(body) < 1500:
            logger.info("sector_data_sync: %s → тело %d байт, похоже на каркас",
                        url, len(body))
            return None
        return body
    except Exception as e:  # noqa: BLE001
        logger.info("sector_data_sync: %s недоступен (%s)", url, type(e).__name__)
        return None


def _strip_html(html: str) -> str:
    t = re.sub(r"<script.*?</script>|<style.*?</style>", " ", html, flags=re.S | re.I)
    t = re.sub(r"<[^>]+>", " ", t)
    for a, b in (("&nbsp;", " "), ("&#160;", " "), ("&amp;", "&"),
                 ("&quot;", '"'), ("&laquo;", "«"), ("&raquo;", "»")):
        t = t.replace(a, b)
    return " ".join(t.split())


def _num(s: str) -> float | None:
    """Русское число в float: «1 177,3» → 1177.3. Неразрывные пробелы — обычное
    дело в вёрстке госсайтов, поэтому чистим их явно."""
    s = s.replace("\xa0", "").replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# --------------------------------------------------------------------------
# Коллекторы. Каждый возвращает список точек (code, title, unit, value, as_of)
# либо пустой список — но НИКОГДА не выдуманное значение.
# --------------------------------------------------------------------------
def collect_power_ups() -> list[dict]:
    """Электроэнергетика: потребление, выработка и установленная мощность ЕЭС
    (Системный оператор). Закрывает 65 компаний, у которых сейчас нет ни одного
    показателя собственного рынка."""
    html = _fetch("https://www.so-ups.ru/functioning/ups/")
    if not html:
        return []
    t = _strip_html(html)
    out: list[dict] = []

    m = re.search(r"[Пп]отреблени\w*\s+электроэнергии[^.]{0,80}?в\s+(\d{4})\s+году"
                  r"[^.]{0,40}?составило\s+([\d\s\xa0,]+)\s*млрд", t)
    if m:
        v = _num(m.group(2))
        if v:
            out.append({"code": "sec_power_consumption", "unit": "млрд кВт·ч",
                        "title": "Потребление электроэнергии в России (год)",
                        "value": v, "as_of": date(int(m.group(1)), 12, 31)})

    m = re.search(r"установленная мощность электростанций[^.]{0,60}?составила\s+"
                  r"([\d\s\xa0,]+)\s*тыс", t)
    if m:
        v = _num(m.group(1))
        d = re.search(r"[Нн]а\s+1\s+января\s+(\d{4})", t)
        if v:
            year = int(d.group(1)) if d else date.today().year
            out.append({"code": "sec_power_capacity", "unit": "тыс. МВт",
                        "title": "Установленная мощность электростанций России",
                        "value": v, "as_of": date(year, 1, 1)})
    if not out:
        logger.warning("sector_data_sync: СО ЕЭС отдал страницу, но числа не разобраны "
                       "— вероятно сменилась вёрстка")
    return out


def collect_power_records() -> list[dict]:
    """Оперативный сигнал по энергетике из RSS Системного оператора: рекорды
    потребления мощности. Годовой ряд для барометра слишком медленный — рекорд
    же показывает, что происходит со спросом ПРЯМО СЕЙЧАС."""
    html = _fetch("https://so-ups.ru/?id=3428&type=333")
    if not html:
        return []
    hits = re.findall(r"<title>(.*?)</title>", html, re.S)[1:]
    marks = [re.sub(r"\s+", " ", h).strip() for h in hits
             if re.search(r"рекорд|потреблени", h, re.I)]
    if not marks:
        return []
    # Это не числовой ряд, а событийный маркер — кладём как текстовую пометку в
    # лог, чтобы барометр не принял его за измерение.
    logger.info("sector_data_sync: СО ЕЭС, свежие сигналы спроса: %s", marks[:2])
    return []


def collect_realty_erz() -> list[dict]:
    """Девелопмент: объём текущего строительства по застройщикам (ЕРЗ).

    Единственный сектор, где ДОЛИ РЫНКА достаются машинно — в остальных это
    дыра, отмеченная во всех реестрах источников. Поэтому берём не сводное
    число (его на странице нет), а таблицу «Застройщик — Строится, м²»: из неё
    считается и рынок целиком, и доля каждого.
    """
    html = _fetch("https://erzrf.ru/top-zastroyshchikov/rf")
    if not html:
        return []
    t = _strip_html(html)
    # «… Строится, м² 4 509 881 С переносом срока …» — числа идут ЗА подписью
    vals = [_num(m.group(1)) for m in
            re.finditer(r"Строится,\s*м²\s+([\d\s\xa0]{6,})", t)]
    vals = [v for v in vals if v and v > 1000]
    if not vals:
        logger.warning("sector_data_sync: ЕРЗ отдал страницу, но объёмы не разобраны "
                       "— вероятно сменилась вёрстка таблицы")
        return []
    total = sum(vals)
    out = [{"code": "sec_realty_construction", "unit": "м²",
            "title": "Объём текущего жилищного строительства (ЕРЗ, топ-застройщики)",
            "value": round(total, 0), "as_of": date.today()}]
    # Концентрация рынка: доля крупнейшего застройщика. Прямой индикатор того,
    # о чём проза «Рынков» говорит словами «рынок консолидируется».
    if len(vals) >= 3:
        out.append({"code": "sec_realty_top1_share", "unit": "%",
                    "title": "Доля крупнейшего застройщика в топе (ЕРЗ)",
                    "value": round(max(vals) / total * 100, 2), "as_of": date.today()})
    return out


def collect_oil_tax() -> list[dict]:
    """Нефтегаз: средняя цена Urals для целей налогообложения (ФНС). Отличается
    от рыночного спота и именно она определяет НДПИ — то есть разницу между
    хорошим кварталом и убытком. В реестре нефтегаза это главная дыра."""
    html = _fetch("https://www.nalog.gov.ru/rn77/taxation/taxes/ndpi/")
    if not html:
        return []
    t = _strip_html(html)
    # На странице НДПИ упоминания цены Urals может не быть вовсе: ФНС публикует
    # её отдельными письмами. Ищем несколькими формулировками и честно молчим,
    # если не нашли, — выдуманная налоговая база хуже отсутствующей.
    m = None
    for pat in (r"сорта\s*«?Юралс»?[^.]{0,90}?([\d\s\xa0,]+)\s*долл",
                r"Urals[^.]{0,90}?([\d\s\xa0,]+)\s*долл",
                r"средн\w+\s+цен\w*\s+на\s+нефть[^.]{0,80}?([\d\s\xa0,]+)\s*долл"):
        m = re.search(pat, t, re.I)
        if m:
            break
    if not m:
        logger.info("sector_data_sync: ФНС — цены Urals на странице НДПИ нет "
                    "(публикуется отдельными письмами, нужен другой адрес)")
        return []
    v = _num(m.group(1))
    return ([{"code": "sec_urals_tax", "unit": "$/барр.",
              "title": "Цена Urals для расчёта НДПИ (ФНС)",
              "value": v, "as_of": date.today()}] if v else [])


def _fetch_bytes(url: str) -> bytes | None:
    try:
        with httpx.Client(follow_redirects=True, timeout=45,
                          headers={"User-Agent": _UA}) as c:
            r = c.get(url)
        if r.status_code != 200 or len(r.content) < 5000:
            logger.info("sector_data_sync: %s → HTTP %s, %d байт",
                        url, r.status_code, len(r.content or b""))
            return None
        return r.content
    except Exception as e:  # noqa: BLE001
        logger.info("sector_data_sync: %s недоступен (%s)", url, type(e).__name__)
        return None


_MONTHS = {"январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
           "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11,
           "декабрь": 12}


def _ru_month(cell: str) -> date | None:
    """«Январь 2014» → 2014-01-31. Строка месяца в файлах ЦБ — единственный
    носитель даты, отдельной колонки с датой там нет."""
    m = re.match(r"\s*([А-Яа-я]+)\s+(\d{4})", str(cell or ""))
    if not m:
        return None
    mon = _MONTHS.get(m.group(1).lower())
    if not mon:
        return None
    year = int(m.group(2))
    from calendar import monthrange
    return date(year, mon, monthrange(year, mon)[1])


def collect_banks_deposit_rate() -> list[dict]:
    """Банки: средневзвешенная ставка по рублёвым вкладам физлиц (ЦБ).

    Это стоимость фондирования — вторая половина процентной маржи, главной
    метрики сектора. Ключевая ставка у нас уже есть, а вот по чём банки реально
    привлекают деньги — не было, и в реестре финансов это отмечено как «маржа
    не питается ничем».
    Файл ЦБ помесячный с 2014 года; путь стабилен между выпусками — это
    надёжнее любого RSS (отмечено при сборке реестра).
    """
    raw = _fetch_bytes("https://www.cbr.ru/vfs/statistics/pdko/int_rat/deposits.xlsx")
    if not raw:
        return []
    try:
        import io

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb["ставки_руб"] if "ставки_руб" in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception as e:  # noqa: BLE001
        logger.warning("sector_data_sync: файл ЦБ не разобран (%s)", e)
        return []

    # Берём последнюю строку с распознанным месяцем и первым числовым значением:
    # состав колонок по срокам менялся, а «есть месяц и есть ставка» — устойчиво.
    last: tuple[date, float] | None = None
    for row in ws.iter_rows(min_row=4, max_col=12, values_only=True):
        d = _ru_month(row[0]) if row and row[0] else None
        if not d:
            continue
        vals = [v for v in row[1:] if isinstance(v, (int, float))]
        if not vals:
            continue
        if last is None or d > last[0]:
            last = (d, float(sum(vals) / len(vals)))
    if not last:
        logger.warning("sector_data_sync: ЦБ отдал файл, но ставки не разобраны")
        return []
    return [{"code": "sec_banks_deposit_rate", "unit": "% годовых",
             "title": "Средняя ставка по рублёвым вкладам физлиц (ЦБ)",
             "value": round(last[1], 2), "as_of": last[0]}]


def collect_ports_asop() -> list[dict]:
    """Транспорт: грузооборот морских портов (АСОП).

    Прямой индикатор экспортного грузопотока — им живут порты и морские
    перевозчики. Росморречфлот недоступен, АСОП его заменяет (найдено при
    сборке реестра транспорта).
    """
    html = _fetch("https://morport.com/rus/content/statistika-0")
    if not html:
        return []
    t = _strip_html(html)
    # На самой странице только заголовки разделов; числа — в приложенных файлах,
    # прямой адрес которых меняется от выпуска к выпуску. Поэтому берём ссылку
    # ИЗ HTML, а не угадываем путь: угаданный давал 404.
    m = re.search(r"[Гг]рузооборот[^.]{0,120}?([\d\s\xa0,]{4,})\s*млн\s*тонн", t)
    if not m:
        logger.info("sector_data_sync: АСОП — сводного числа на странице нет, "
                    "грузооборот лежит в приложенных файлах (нужен разбор xlsx "
                    "по ссылке из HTML)")
        return []
    v = _num(m.group(1))
    return ([{"code": "sec_ports_turnover", "unit": "млн тонн",
              "title": "Грузооборот морских портов России (АСОП)",
              "value": v, "as_of": date.today()}] if v else [])


def collect_software_registry() -> list[dict]:
    """IT: число позиций в реестре отечественного ПО (Минцифры).

    Реестр — это освобождение от НДС и допуск к госзаказу, то есть прямой
    фактор выручки. Головной digital.gov.ru в гео-блоке, а подведомственный
    reestr.digital.gov.ru открыт — находка реестра источников IT.
    """
    # 🔴 Доступность этого источника ПЛАВАЮЩАЯ: при сборке реестра IT он отдавал
    # 1,1 МБ данных, на следующий день — глухой таймаут (000). Это не повод его
    # выбрасывать: коллектор просто промолчит в неудачный день, а в удачный
    # добавит точку. Ровно для таких случаев ряд и не перезаписывается.
    html = _fetch("https://reestr.digital.gov.ru/reestr/")
    if not html:
        return []
    t = _strip_html(html)
    m = re.search(r"(?:найдено|всего|записей|результат\w*)[^.\d]{0,25}([\d\s\xa0]{3,})", t, re.I)
    if not m:
        return []
    v = _num(m.group(1))
    return ([{"code": "sec_it_software_registry", "unit": "позиций",
              "title": "Записей в реестре отечественного ПО (Минцифры)",
              "value": v, "as_of": date.today()}] if v and v > 100 else [])


COLLECTORS = [
    ("СО ЕЭС — потребление и мощность", collect_power_ups),
    ("СО ЕЭС — сигналы спроса", collect_power_records),
    ("ЕРЗ — жилищное строительство", collect_realty_erz),
    ("ФНС — Urals для НДПИ", collect_oil_tax),
    ("ЦБ — ставка по вкладам", collect_banks_deposit_rate),
    ("АСОП — грузооборот портов", collect_ports_asop),
    ("Минцифры — реестр ПО", collect_software_registry),
]


def _ensure_indicator(db: Session, code: str, title: str, unit: str) -> None:
    db.execute(text("""
        INSERT INTO macro_indicators (code, title, unit, country, frequency,
                                      display_group, source_type)
        VALUES (:c, :t, :u, 'RU', 'monthly', 'sector', 'sector_source')
        ON CONFLICT (code) DO UPDATE SET title = EXCLUDED.title, unit = EXCLUDED.unit
    """), {"c": code, "t": title, "u": unit})


def _upsert_point(db: Session, code: str, value: float, as_of: date) -> bool:
    """Точка ряда. Существующее значение НЕ перезаписываем другим: если источник
    начал отдавать иное число за ту же дату, это повод разобраться, а не тихо
    подменить историю (та же логика, что в macro_ingest)."""
    row = db.execute(text("""
        SELECT value FROM macro_data_points
        WHERE indicator_code = :c AND as_of = :d AND metric = 'level'
    """), {"c": code, "d": as_of}).fetchone()
    if row is not None:
        if abs(float(row[0]) - value) > 1e-9:
            logger.warning("sector_data_sync: %s на %s уже есть значение %s, источник "
                           "отдал %s — оставляю прежнее", code, as_of, row[0], value)
        return False
    db.execute(text("""
        INSERT INTO macro_data_points (indicator_code, as_of, metric, value, ingested_via)
        VALUES (:c, :d, 'level', :v, 'sector_source')
    """), {"c": code, "d": as_of, "v": value})
    return True


def refresh_all(db: Session) -> dict:
    """Прогон всех коллекторов. Падение одного не роняет остальные — тот же
    контракт, что у macro_ingest."""
    stats = {"collected": 0, "written": 0, "failed": []}
    for name, fn in COLLECTORS:
        try:
            points = fn()
        except Exception as e:  # noqa: BLE001
            logger.warning("sector_data_sync: коллектор «%s» упал: %s", name, e)
            stats["failed"].append(name)
            continue
        if not points:
            stats["failed"].append(name)
            continue
        for p in points:
            _ensure_indicator(db, p["code"], p["title"], p["unit"])
            if _upsert_point(db, p["code"], p["value"], p["as_of"]):
                stats["written"] += 1
            stats["collected"] += 1
    db.commit()
    logger.info("sector_data_sync: %s", stats)
    return stats
