"""Синхронизация МЕСЯЧНЫХ мировых цен сырья на платформе (Направление:
«Товар компании», commodity_exposure в market.json компаний, см.
.claude/agents/market-analyst.md ОБНОВЛЕНИЕ v6).

🔴 2026-07-22: FORTS-ряды (золото/серебро/платина/медь/газ/нефть как
российские деривативы через instrument_history) НЕ дают многолетней
глубины на практике — контракт живёт ~3 месяца, история истёкших
контрактов не бэкафиллится (см. docs/status.md, «непрерывные ряды»).
Поэтому здесь дублируем WB Pink Sheet и по этим товарам тоже — как честную
альтернативу с реальной глубиной (2016+, помесячно), а не полагаемся на
теоретически «живой», но фактически короткий FORTS-канал. Urals — отдельно,
через macro_tankermap_sync.py (ДНЕВНОЙ, реально глубокий с 2024).
Алюминий/никель/уголь/руда/удобрения на MOEX вообще не торгуются — только
WB. Здесь берём официальный МЕСЯЧНЫЙ бенчмарк Всемирного банка (World Bank
Commodity Price Data, «Pink Sheet»), обновляется раз в месяц, публикуется
с 1960 года, свободный доступ без ключа/авторизации.

Источник — НЕ прямая CSV-ссылка (адрес XLSX-файла меняется, судя по всему,
ежегодно —ハッシュ-префикс в URL другой у 2025 и 2026 годовых серий
документов thedocs.worldbank.org), поэтому дискавери: страница
worldbank.org/en/research/commodity-markets стабильно содержит прямую
ссылку на ТЕКУЩИЙ файл CMO-Historical-Data-Monthly.xlsx — регексом достаём
её оттуда при каждом синке, не хардкодим адрес файла.

Лист "Monthly Prices": строка 5 — названия колонок, строка 6 — единицы,
дальше строки вида ('2026M06', <цена1>, <цена2>, ...) по одной в месяц.
Курируемый маппинг колонка → indicator_code (см. _COLUMNS ниже) — не берём
всё подряд (89 колонок в файле, нам нужны только сырьевые входы/выходы без
своего биржевого ряда).
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date

import httpx
import openpyxl
from sqlalchemy.orm import Session

from app.services.macro_ingest import upsert_point

logger = logging.getLogger(__name__)

_HTTP = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"}
_LANDING = "https://www.worldbank.org/en/research/commodity-markets"
_XLSX_RE = re.compile(
    r'https://thedocs\.worldbank\.org/en/doc/[a-z0-9]+-\d+/related/CMO-Historical-Data-Monthly\.xlsx')

# Название колонки в файле (строка 5 листа "Monthly Prices") → наш indicator_code.
# Курируемый список — не берём все 71+ колонку файла, только сырьё, реально
# нужное для commodity_exposure карточек компаний платформы (плюс какао/сахар —
# закрывает пробел benchmark_status:"none" у кондитерских разборов).
_COLUMNS = {
    "Aluminum": "wb_aluminum",
    "Nickel": "wb_nickel",
    "Coal, Australian": "wb_coal",
    "Iron ore, cfr spot": "wb_iron_ore",
    "Phosphate rock": "wb_phosphate_rock",
    "DAP": "wb_dap",
    "Potassium chloride **": "wb_potash",
    "Urea ": "wb_urea",  # в файле с хвостовым пробелом, см. заголовок
    "Crude oil, Brent": "wb_oil_brent",
    "Crude oil, WTI": "wb_oil_wti",
    "Natural gas, US": "wb_gas_us",
    "Natural gas, Europe": "wb_gas_europe",
    "Liquefied natural gas, Japan": "wb_gas_lng_japan",
    "Copper": "wb_copper",
    "Lead": "wb_lead",
    "Tin": "wb_tin",
    "Zinc": "wb_zinc",
    "Gold": "wb_gold",
    "Platinum": "wb_platinum",
    "Silver": "wb_silver",
    "Cocoa": "wb_cocoa",
    "Sugar, world": "wb_sugar_world",
    "Wheat, US HRW": "wb_wheat",
    "Fish meal": "wb_fish_meal",
    "Soybean meal": "wb_soybean_meal",
}
_UNITS = {
    "wb_aluminum": "usd/mt", "wb_nickel": "usd/mt", "wb_coal": "usd/mt",
    "wb_iron_ore": "usd/dmtu", "wb_phosphate_rock": "usd/mt", "wb_dap": "usd/mt",
    "wb_potash": "usd/mt", "wb_urea": "usd/mt",
    "wb_oil_brent": "usd/bbl", "wb_oil_wti": "usd/bbl",
    "wb_gas_us": "usd/mmbtu", "wb_gas_europe": "usd/mmbtu", "wb_gas_lng_japan": "usd/mmbtu",
    "wb_copper": "usd/mt", "wb_lead": "usd/mt", "wb_tin": "usd/mt", "wb_zinc": "usd/mt",
    "wb_gold": "usd/oz", "wb_platinum": "usd/oz", "wb_silver": "usd/oz",
    "wb_cocoa": "usd/kg", "wb_sugar_world": "usd/kg", "wb_wheat": "usd/mt",
    "wb_fish_meal": "usd/mt", "wb_soybean_meal": "usd/mt",
}
_PERIOD_RE = re.compile(r"^(\d{4})M(\d{1,2})$")


def _discover_xlsx_url() -> str | None:
    try:
        r = httpx.get(_LANDING, timeout=25, headers=_HTTP, follow_redirects=True)
        r.raise_for_status()
    except Exception as e:  # noqa: BLE001
        logger.warning("WB Pink Sheet: страница-дискавери недоступна: %s", type(e).__name__)
        return None
    m = _XLSX_RE.search(r.text)
    return m.group(0) if m else None


def _month_end(year: int, month: int) -> date:
    from calendar import monthrange
    return date(year, month, monthrange(year, month)[1])


def sync_wb_commodities(db: Session, months_back: int = 3) -> dict:
    """months_back=N — сколько последних месяцев ряда обновить (файл целиком
    содержит историю с 1960, но нам нужен только свежий хвост — старое уже
    стабильно и не меняется задним числом). Суточный крон — months_back=3
    (догоняет пропуски/ревизии), разовый бэкфилл истории — вызвать отдельно
    с months_back побольше (например 120 = 10 лет)."""
    url = _discover_xlsx_url()
    if not url:
        return {"error": "discovery_failed"}
    try:
        r = httpx.get(url, timeout=60, headers=_HTTP, follow_redirects=True)
        r.raise_for_status()
    except Exception as e:  # noqa: BLE001
        logger.warning("WB Pink Sheet: файл недоступен: %s", type(e).__name__)
        return {"error": f"fetch_failed:{type(e).__name__}", "url": url}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        ws = wb["Monthly Prices"]
        headers = next(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    except Exception as e:  # noqa: BLE001
        logger.warning("WB Pink Sheet: не удалось разобрать XLSX: %s", type(e).__name__)
        return {"error": f"parse_failed:{type(e).__name__}", "url": url}

    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in _COLUMNS:
            col_idx[_COLUMNS[h]] = i
    missing = set(_COLUMNS.values()) - set(col_idx.keys())
    if missing:
        logger.warning("WB Pink Sheet: не нашли колонки в файле: %s (формат файла мог измениться)", missing)

    rows = list(ws.iter_rows(min_row=7, values_only=True))  # с 7 строки — данные (5=заголовки, 6=единицы)
    rows = [r for r in rows if r and r[0] and _PERIOD_RE.match(str(r[0]))]
    todo = rows[-months_back:] if months_back < len(rows) else rows

    saved, skipped = 0, 0
    for row in todo:
        m = _PERIOD_RE.match(str(row[0]))
        year, month = int(m.group(1)), int(m.group(2))
        d = _month_end(year, month)
        for code, idx in col_idx.items():
            val = row[idx] if idx < len(row) else None
            if val is None or not isinstance(val, (int, float)):
                continue
            res = upsert_point(db, code, d, "level", float(val), unit=_UNITS.get(code),
                               source="World Bank Commodity Price Data (Pink Sheet)",
                               source_url=url, ingested_via="wb", commit=False)
            if res in ("insert", "revise"):
                saved += 1
            else:
                skipped += 1
    db.commit()
    logger.info("WB Pink Sheet: %d сохранено, %d без изменений (за %d последних месяцев, %d колонок)",
                saved, skipped, len(todo), len(col_idx))
    return {"months": len(todo), "columns": list(col_idx.keys()), "missing_columns": list(missing),
            "saved": saved, "skipped": skipped, "url": url}
